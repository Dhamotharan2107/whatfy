#!/usr/bin/env python3
"""
MCP Server for WhatsApp Message Service
Supports Excel uploads, OAuth authentication, and rate limiting
"""

import asyncio
import json
import logging
import os
import secrets
import time
import uuid
from datetime import datetime, timedelta
from typing import Any, AsyncIterator, Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, HTTPException, status
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from pydantic import BaseModel
import uvicorn
from openpyxl import load_workbook
from collections import defaultdict
import threading

try:
    import qrcode as _qrlib
    QR_OK = True
except ImportError:
    QR_OK = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("MCP_SERVER")

# ── Constants ──────────────────────────────────────────────────────────────────────
RATE_LIMIT_WINDOW = 60  # seconds
RATE_LIMIT_REQUESTS = 10  # requests per window per user
RATE_LIMIT_BURST = 3  # burst requests per window
MAX_EXCEL_SIZE = 10 * 1024 * 1024  # 10MB
MAX_PHONES_PER_FILE = 5000
MAX_BATCH_SIZE = 1000

# ── OAuth Configuration ────────────────────────────────────────────────────────────
class OAuthConfig:
    """User-based OAuth configuration"""
    def __init__(self, client_id: str, client_secret: str, grant_type: str, scope: str, redirect_uri: str = ""):
        self.client_id = client_id
        self.client_secret = client_secret
        self.grant_type = grant_type
        self.scope = scope
        self.redirect_uri = redirect_uri

    def to_dict(self) -> dict:
        return {
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "grant_type": self.grant_type,
            "scope": self.scope,
            "redirect_uri": self.redirect_uri
        }

# Global OAuth configs storage (user_id -> OAuthConfig)
oauth_configs: dict[str, OAuthConfig] = {}
oauth_lock = threading.Lock()

def generate_client_secret() -> str:
    """Generate a secure random client secret"""
    return secrets.token_urlsafe(32)

# ── Rate Limiter ──────────────────────────────────────────────────────────────────
class RateLimiter:
    """In-memory rate limiter with sliding window algorithm"""
    
    def __init__(self):
        self.requests: dict[str, list[float]] = defaultdict(list)
        self.lock = threading.RLock()
    
    async def check_rate_limit(self, user_id: str) -> dict[str, Any]:
        """Check if request is within rate limits. Returns rate limit info and HTTP status."""
        current_time = time.time()
        
        with self.lock:
            window_start = current_time - RATE_LIMIT_WINDOW
            
            # Clean up old requests
            self.requests[user_id] = [
                ts for ts in self.requests[user_id]
                if ts > window_start
            ]
            
            request_count = len(self.requests[user_id])
            
            # Check rate limit
            if request_count >= RATE_LIMIT_REQUESTS:
                # Calculate retry after time
                oldest_request = self.requests[user_id][0]
                retry_after = int(oldest_request - window_start) + 1
                
                return {
                    "allowed": False,
                    "code": 429,
                    "message": f"Rate limit exceeded. {retry_after} seconds until next request.",
                    "retry_after": retry_after,
                    "current_requests": request_count,
                    "limit": RATE_LIMIT_REQUESTS,
                    "window": RATE_LIMIT_WINDOW
                }
            
            # Check burst limit
            if request_count > RATE_LIMIT_BURST:
                retry_after = int(time.time() - self.requests[user_id][0]) + 1
                return {
                    "allowed": False,
                    "code": 429,
                    "message": f"Burst limit exceeded. {retry_after} seconds until next request.",
                    "retry_after": retry_after,
                    "current_requests": request_count,
                    "limit": RATE_LIMIT_BURST,
                    "window": RATE_LIMIT_WINDOW
                }
            
            # Add current request
            self.requests[user_id].append(current_time)
            
            return {
                "allowed": True,
                "code": 200,
                "message": "Request allowed",
                "current_requests": request_count + 1,
                "limit": RATE_LIMIT_REQUESTS,
                "window": RATE_LIMIT_WINDOW
            }
    
    def reset_user(self, user_id: str):
        """Reset rate limit for a specific user"""
        with self.lock:
            if user_id in self.requests:
                del self.requests[user_id]

# Global rate limiter
rate_limiter = RateLimiter()

# ── Excel Processing ───────────────────────────────────────────────────────────────
def process_excel_file(file_path: str) -> list[str]:
    """Process Excel file and extract phone numbers"""
    phones = []
    
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        
        # Try different sheet names
        sheet_names = wb.sheetnames
        sheet = wb.active  # Default to active sheet
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            # Find phone number in any column
            for value in row:
                if value and isinstance(value, (int, float, str)):
                    phone_str = str(value).strip()
                    
                    # Remove non-digit characters
                    digits = ''.join(c for c in phone_str if c.isdigit())
                    
                    # Check if it's a valid phone number (10+ digits)
                    if len(digits) >= 10:
                        phones.append(digits)
        
        wb.close()
        
        # Limit to maximum phones per file
        phones = phones[:MAX_PHONES_PER_FILE]
        
        return phones
    
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise HTTPException(
            status_code=400,
            detail=f"Failed to process Excel file: {str(e)}"
        )

# ── MCP Server ────────────────────────────────────────────────────────────────────

class MCPResponse:
    """MCP Response wrapper"""
    def __init__(self, content: Any, success: bool = True, metadata: Optional[dict] = None):
        self.content = content
        self.success = success
        self.metadata = metadata or {}
    
    def to_dict(self) -> dict:
        return {
            "content": self.content,
            "success": self.success,
            "metadata": self.metadata,
            "timestamp": datetime.utcnow().isoformat()
        }

class MCPServer:
    """MCP Server implementation"""
    
    def __init__(self, base_url: str = "http://localhost:8000"):
        self.base_url = base_url
        self.mcp_commands = {
            "init": self._handle_init,
            "oauth_register": self._handle_oauth_register,
            "oauth_get_config": self._handle_oauth_get_config,
            "oauth_authorize": self._handle_oauth_authorize,
            "oauth_token": self._handle_oauth_token,
            "upload_excel": self._handle_upload_excel,
            "send_messages": self._handle_send_messages,
            "get_batch_status": self._handle_get_batch_status,
            "get_statistics": self._handle_get_statistics,
        }
    
    async def process_request(self, user_id: str, command: str, params: dict) -> dict:
        """Process MCP command"""
        rate_limit_info = await rate_limiter.check_rate_limit(user_id)
        
        if not rate_limit_info["allowed"]:
            return {
                "success": False,
                "error": rate_limit_info["message"],
                "code": rate_limit_info["code"],
                "retry_after": rate_limit_info.get("retry_after", 0)
            }
        
        handler = self.mcp_commands.get(command)
        if not handler:
            return {
                "success": False,
                "error": f"Unknown command: {command}",
                "code": 400
            }
        
        try:
            result = await handler(user_id, params)
            return {
                "success": True,
                "data": result,
                "metadata": {
                    "rate_limit": {
                        "current": rate_limit_info["current_requests"],
                        "limit": rate_limit_info["limit"],
                        "window": rate_limit_info["window"]
                    }
                }
            }
        except HTTPException as e:
            return {
                "success": False,
                "error": e.detail,
                "code": e.status_code
            }
        except Exception as e:
            logger.error(f"Error in command {command}: {e}")
            return {
                "success": False,
                "error": str(e),
                "code": 500
            }
    
    async def _handle_init(self, user_id: str, params: dict) -> dict:
        """Initialize MCP connection"""
        return {
            "server_version": "1.0.0",
            "features": [
                "oauth_authentication",
                "excel_upload",
                "rate_limiting",
                "batch_sending"
            ],
            "rate_limits": {
                "per_window": RATE_LIMIT_REQUESTS,
                "window_seconds": RATE_LIMIT_WINDOW,
                "burst": RATE_LIMIT_BURST
            }
        }
    
    async def _handle_oauth_register(self, user_id: str, params: dict) -> dict:
        """Register OAuth client credentials for user"""
        grant_type = params.get("grant_type", "client_credentials")
        scope = params.get("scope", "whatsapp_message")
        redirect_uri = params.get("redirect_uri", "")
        
        with oauth_lock:
            client_id = secrets.token_urlsafe(16)
            client_secret = generate_client_secret()
            
            config = OAuthConfig(
                client_id=client_id,
                client_secret=client_secret,
                grant_type=grant_type,
                scope=scope,
                redirect_uri=redirect_uri
            )
            
            oauth_configs[user_id] = config
        
        return {
            "message": "OAuth client registered successfully",
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": grant_type,
            "scope": scope,
            "redirect_uri": redirect_uri
        }
    
    async def _handle_oauth_get_config(self, user_id: str, params: dict) -> dict:
        """Get OAuth configuration for user"""
        with oauth_lock:
            if user_id not in oauth_configs:
                raise HTTPException(
                    status_code=404,
                    detail="OAuth client not found. Register first using oauth_register."
                )
            
            config = oauth_configs[user_id]
            
            # Don't return full secret in GET requests
            return {
                "client_id": config.client_id,
                "grant_type": config.grant_type,
                "scope": config.scope,
                "redirect_uri": config.redirect_uri,
                "has_secret": True
            }
    
    async def _handle_oauth_authorize(self, user_id: str, params: dict) -> dict:
        """Get OAuth authorization URL"""
        with oauth_lock:
            if user_id not in oauth_configs:
                raise HTTPException(
                    status_code=404,
                    detail="OAuth client not found. Register first using oauth_register."
                )
            
            config = oauth_configs[user_id]
            
            auth_url = f"{self.base_url}/oauth/authorize"
            params_to_add = {
                "client_id": config.client_id,
                "response_type": "code",
                "scope": config.scope,
                "redirect_uri": config.redirect_uri
            }
            
            # Add random state for CSRF protection
            state = secrets.token_urlsafe(16)
            params_to_add["state"] = state
            
            return {
                "authorization_url": auth_url,
                "params": params_to_add,
                "state": state,
                "description": "User must visit this URL to authorize the client"
            }
    
    async def _handle_oauth_token(self, user_id: str, params: dict) -> dict:
        """Exchange authorization code for access token"""
        auth_code = params.get("auth_code")
        redirect_uri = params.get("redirect_uri", "")
        
        if not auth_code:
            raise HTTPException(
                status_code=400,
                detail="Authorization code is required"
            )
        
        with oauth_lock:
            if user_id not in oauth_configs:
                raise HTTPException(
                    status_code=404,
                    detail="OAuth client not found. Register first using oauth_register."
                )
            
            config = oauth_configs[user_id]
            
            # Simulate token exchange
            access_token = secrets.token_urlsafe(32)
            token_type = "bearer"
            expires_in = 3600  # 1 hour
            
            # Store token for user
            if not hasattr(config, "tokens"):
                config.tokens = {}
            
            config.tokens[user_id] = {
                "access_token": access_token,
                "token_type": token_type,
                "expires_in": expires_in,
                "expires_at": int(time.time()) + expires_in,
                "refresh_token": secrets.token_urlsafe(32)
            }
        
        return {
            "message": "Token generated successfully",
            "access_token": access_token,
            "token_type": token_type,
            "expires_in": expires_in,
            "expires_at": datetime.fromtimestamp(int(time.time()) + expires_in).isoformat()
        }
    
    async def _handle_upload_excel(self, user_id: str, params: dict) -> dict:
        """Upload Excel file and extract phone numbers"""
        file_path = params.get("file_path")
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(
                status_code=400,
                detail="File path is required and file must exist"
            )
        
        # Check file size
        file_size = os.path.getsize(file_path)
        if file_size > MAX_EXCEL_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File size exceeds {MAX_EXCEL_SIZE // (1024*1024)}MB limit"
            )
        
        # Process Excel file
        phones = process_excel_file(file_path)
        
        if not phones:
            raise HTTPException(
                status_code=400,
                detail="No phone numbers found in Excel file"
            )
        
        # Limit phones per file
        if len(phones) > MAX_PHONES_PER_FILE:
            phones = phones[:MAX_PHONES_PER_FILE]
        
        # Save phones for batch sending
        batch_id = str(uuid.uuid4())
        
        if not hasattr(self, "phone_batches"):
            self.phone_batches = {}
        
        self.phone_batches[batch_id] = {
            "phones": phones,
            "uploaded_at": datetime.utcnow().isoformat(),
            "file_size": file_size
        }
        
        return {
            "batch_id": batch_id,
            "phones": phones,
            "total_phones": len(phones),
            "message": f"Successfully uploaded {len(phones)} phone numbers"
        }
    
    async def _handle_send_messages(self, user_id: str, params: dict) -> dict:
        """Send messages to phone numbers"""
        batch_id = params.get("batch_id")
        message_template = params.get("message_template", "Hi")
        dry_run = params.get("dry_run", False)
        
        if not batch_id:
            raise HTTPException(
                status_code=400,
                detail="Batch ID is required"
            )
        
        if not hasattr(self, "phone_batches") or batch_id not in self.phone_batches:
            raise HTTPException(
                status_code=404,
                detail="Batch not found"
            )
        
        batch = self.phone_batches[batch_id]
        phones = batch["phones"]
        
        if not phones:
            raise HTTPException(
                status_code=400,
                detail="No phones in batch"
            )
        
        # Limit batch size
        if len(phones) > MAX_BATCH_SIZE:
            phones = phones[:MAX_BATCH_SIZE]
        
        # Simulate sending messages
        if not dry_run:
            # Here you would integrate with actual WhatsApp sending
            results = []
            for i, phone in enumerate(phones):
                result = {
                    "phone": phone,
                    "status": "sent",
                    "message_id": f"msg_{uuid.uuid4()}",
                    "timestamp": datetime.utcnow().isoformat()
                }
                results.append(result)
        else:
            results = [
                {
                    "phone": phone,
                    "status": "dry_run",
                    "message_id": f"dry_{uuid.uuid4()}",
                    "timestamp": datetime.utcnow().isoformat()
                }
                for phone in phones
            ]
        
        return {
            "batch_id": batch_id,
            "total_phones": len(phones),
            "sent": len(results),
            "results": results[:100],  # Limit results to prevent huge responses
            "message": f"Message template '{message_template}' would be sent to {len(phones)} phone numbers"
        }
    
    async def _handle_get_batch_status(self, user_id: str, params: dict) -> dict:
        """Get batch status"""
        batch_id = params.get("batch_id")
        
        if not batch_id:
            raise HTTPException(
                status_code=400,
                detail="Batch ID is required"
            )
        
        if not hasattr(self, "phone_batches") or batch_id not in self.phone_batches:
            raise HTTPException(
                status_code=404,
                detail="Batch not found"
            )
        
        batch = self.phone_batches[batch_id]
        
        return {
            "batch_id": batch_id,
            "phones_count": len(batch["phones"]),
            "uploaded_at": batch["uploaded_at"],
            "file_size": batch["file_size"]
        }
    
    async def _handle_get_statistics(self, user_id: str, params: dict) -> dict:
        """Get statistics"""
        return {
            "total_batches": len(self.phone_batches) if hasattr(self, "phone_batches") else 0,
            "total_phones_processed": sum(
                batch["phones_count"] 
                for batch in self.phone_batches.values() 
                if hasattr(batch, "phones_count")
            ) if hasattr(self, "phone_batches") else 0,
            "oauth_clients_registered": len(oauth_configs),
            "rate_limits": {
                "window": RATE_LIMIT_WINDOW,
                "requests_per_window": RATE_LIMIT_REQUESTS,
                "burst": RATE_LIMIT_BURST
            }
        }

# ── FastAPI Application ────────────────────────────────────────────────────────────

mcp_server = MCPServer()

app = FastAPI(
    title="MCP WhatsApp Server",
    description="MCP Server for WhatsApp messaging with OAuth, Excel upload, and rate limiting",
    version="1.0.0"
)

# Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)

class MessageTemplateRequest(BaseModel):
    """Request model for sending messages"""
    batch_id: str
    message_template: str = "Hi"
    dry_run: bool = False

class ExcelUploadRequest(BaseModel):
    """Request model for Excel upload"""
    file_path: str

class BatchStatusRequest(BaseModel):
    """Request model for batch status"""
    batch_id: str

@app.get("/")
async def root():
    """Root endpoint - MCP server info"""
    return {
        "name": "MCP WhatsApp Server",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "oauth_register": "POST /api/oauth/register",
            "oauth_authorize": "POST /api/oauth/authorize",
            "oauth_token": "POST /api/oauth/token",
            "upload_excel": "POST /api/upload/excel",
            "send_messages": "POST /api/messages/send",
            "batch_status": "GET /api/batch/{batch_id}",
            "statistics": "GET /api/statistics"
        }
    }

@app.post("/api/oauth/register")
async def oauth_register(
    user_id: str,
    grant_type: str = "client_credentials",
    scope: str = "whatsapp_message",
    redirect_uri: str = ""
):
    """Register OAuth client credentials for user"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="oauth_register",
        params={
            "grant_type": grant_type,
            "scope": scope,
            "redirect_uri": redirect_uri
        }
    )
    return result

@app.get("/api/oauth/config")
async def oauth_get_config(user_id: str):
    """Get OAuth configuration for user"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="oauth_get_config",
        params={}
    )
    return result

@app.post("/api/oauth/authorize")
async def oauth_authorize(user_id: str, redirect_uri: str = ""):
    """Get OAuth authorization URL"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="oauth_authorize",
        params={"redirect_uri": redirect_uri}
    )
    return result

@app.post("/api/oauth/token")
async def oauth_token(user_id: str, auth_code: str, redirect_uri: str = ""):
    """Exchange authorization code for access token"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="oauth_token",
        params={
            "auth_code": auth_code,
            "redirect_uri": redirect_uri
        }
    )
    return result

@app.post("/api/upload/excel")
async def upload_excel(user_id: str, file_path: str):
    """Upload Excel file and extract phone numbers"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="upload_excel",
        params={"file_path": file_path}
    )
    return result

@app.post("/api/messages/send")
async def send_messages(
    user_id: str,
    batch_id: str,
    message_template: str = "Hi",
    dry_run: bool = False
):
    """Send messages to phone numbers"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="send_messages",
        params={
            "batch_id": batch_id,
            "message_template": message_template,
            "dry_run": dry_run
        }
    )
    return result

@app.get("/api/batch/{batch_id}")
async def get_batch_status(user_id: str, batch_id: str):
    """Get batch status"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="get_batch_status",
        params={"batch_id": batch_id}
    )
    return result

@app.get("/api/statistics")
async def get_statistics(user_id: str):
    """Get statistics"""
    result = await mcp_server.process_request(
        user_id=user_id,
        command="get_statistics",
        params={}
    )
    return result

@app.get("/api/rate-limit/check")
async def check_rate_limit(user_id: str):
    """Check current rate limit status"""
    rate_limit_info = await rate_limiter.check_rate_limit(user_id)
    return rate_limit_info

@app.delete("/api/rate-limit/reset/{user_id}")
async def reset_rate_limit(user_id: str):
    """Reset rate limit for user (admin function)"""
    rate_limiter.reset_user(user_id)
    return {
        "success": True,
        "message": f"Rate limit reset for user {user_id}"
    }

# ── Main Entry Point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    
    port = int(os.environ.get("MCP_PORT", "8000"))
    host = os.environ.get("MCP_HOST", "0.0.0.0")
    
    logger.info(f"Starting MCP WhatsApp Server on {host}:{port}")
    uvicorn.run(
        app,
        host=host,
        port=port,
        log_level="info"
    )
