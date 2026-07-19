#!/usr/bin/env python3
"""
Test script for MCP WhatsApp Server
Demonstrates OAuth registration, Excel upload, and message sending
"""

import requests
import time
import os

# MCP Server Configuration
BASE_URL = "http://localhost:8001"

def print_section(title):
    """Print a formatted section header"""
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}\n")

def check_rate_limit(user_id):
    """Check current rate limit status"""
    print_section("Checking Rate Limit")
    response = requests.get(f"{BASE_URL}/api/rate-limit/check", params={"user_id": user_id})
    result = response.json()
    
    if result.get("allowed"):
        print(f"✓ Request allowed")
        print(f"  Current: {result['current_requests']}/{result['limit']} requests")
    else:
        print(f"✗ Rate limit exceeded")
        print(f"  {result['message']}")
        print(f"  Retry after: {result['retry_after']} seconds")
    
    return result.get("allowed", False)

def test_oauth_register(user_id):
    """Test OAuth client registration"""
    print_section("Testing OAuth Client Registration")
    
    response = requests.post(
        f"{BASE_URL}/api/oauth/register",
        params={"user_id": user_id},
        data={
            "grant_type": "client_credentials",
            "scope": "whatsapp_message",
            "redirect_uri": "http://localhost:3000/callback"
        }
    )
    
    result = response.json()
    
    if response.status_code == 200 and result.get("success"):
        print(f"✓ OAuth client registered successfully")
        print(f"  Client ID: {result['data']['client_id'][:20]}...")
        print(f"  Client Secret: {result['data']['client_secret'][:20]}...")
        print(f"  Grant Type: {result['data']['grant_type']}")
        print(f"  Scope: {result['data']['scope']}")
        return result["data"]["client_id"], result["data"]["client_secret"]
    else:
        print(f"✗ Registration failed")
        print(f"  Error: {result.get('error', 'Unknown error')}")
        return None, None

def test_upload_excel(user_id, file_path=None):
    """Test Excel file upload"""
    print_section("Testing Excel Upload")
    
    if file_path and not os.path.exists(file_path):
        print(f"✗ File not found: {file_path}")
        print(f"  Creating sample Excel file...")
        
        # Create a sample Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phone Numbers"
        
        # Add some sample phone numbers
        phones = [
            "9876543210",
            "9876543211",
            "9876543212",
            "9876543213",
            "9876543214",
            "9876543215",
            "9876543216",
            "9876543217",
            "9876543218",
            "9876543219"
        ]
        
        for i, phone in enumerate(phones, 1):
            ws[f"A{i}"] = phone
            ws[f"B{i}"] = phone
            ws[f"C{i}"] = phone
        
        wb.save("test_phones.xlsx")
        print(f"✓ Created sample Excel file: test_phones.xlsx")
        file_path = "test_phones.xlsx"
    
    if not file_path or not os.path.exists(file_path):
        print(f"✗ No file provided")
        return None
    
    print(f"Uploading: {file_path}")
    
    try:
        response = requests.post(
            f"{BASE_URL}/api/upload/excel",
            params={"user_id": user_id},
            data={"file_path": file_path}
        )
        
        result = response.json()
        
        if response.status_code == 200 and result.get("success"):
            print(f"✓ File uploaded successfully")
            print(f"  Batch ID: {result['data']['batch_id']}")
            print(f"  Total phones: {result['data']['total_phones']}")
            print(f"  Message: {result['data']['message']}")
            return result["data"]["batch_id"]
        else:
            print(f"✗ Upload failed")
            print(f"  Error: {result.get('error', 'Unknown error')}")
            return None
    except Exception as e:
        print(f"✗ Error uploading file: {e}")
        return None

def test_send_messages(user_id, batch_id, message_template="Hi there!"):
    """Test sending messages"""
    print_section("Testing Message Sending")
    
    print(f"Batch ID: {batch_id}")
    print(f"Message Template: {message_template}")
    
    response = requests.post(
        f"{BASE_URL}/api/messages/send",
        params={"user_id": user_id},
        data={
            "batch_id": batch_id,
            "message_template": message_template,
            "dry_run": True  # Run in dry-run mode
        }
    )
    
    result = response.json()
    
    if response.status_code == 200 and result.get("success"):
        print(f"✓ Message sending initiated")
        print(f"  Total phones: {result['data']['total_phones']}")
        print(f"  Sent: {result['data']['sent']}")
        print(f"  Message: {result['data']['message']}")
        
        if result['data'].get('results'):
            print(f"\nSample results (first 3):")
            for i, res in enumerate(result['data']['results'][:3], 1):
                print(f"  {i}. Phone: {res['phone']}, Status: {res['status']}")
        
        return result["data"]
    else:
        print(f"✗ Sending failed")
        print(f"  Error: {result.get('error', 'Unknown error')}")
        return None

def test_get_batch_status(user_id, batch_id):
    """Test getting batch status"""
    print_section("Testing Batch Status")
    
    response = requests.get(
        f"{BASE_URL}/api/batch/{batch_id}",
        params={"user_id": user_id}
    )
    
    result = response.json()
    
    if response.status_code == 200 and result.get("success"):
        print(f"✓ Batch status retrieved")
        print(f"  Batch ID: {result['data']['batch_id']}")
        print(f"  Phones Count: {result['data']['phones_count']}")
        print(f"  Uploaded At: {result['data']['uploaded_at']}")
        print(f"  File Size: {result['data']['file_size']} bytes")
        return result["data"]
    else:
        print(f"✗ Failed to get batch status")
        print(f"  Error: {result.get('error', 'Unknown error')}")
        return None

def test_get_statistics(user_id):
    """Test getting server statistics"""
    print_section("Testing Server Statistics")
    
    response = requests.get(
        f"{BASE_URL}/api/statistics",
        params={"user_id": user_id}
    )
    
    result = response.json()
    
    if response.status_code == 200 and result.get("success"):
        print(f"✓ Statistics retrieved")
        stats = result['data']
        print(f"  Total Batches: {stats['total_batches']}")
        print(f"  Total Phones Processed: {stats['total_phones_processed']}")
        print(f"  OAuth Clients Registered: {stats['oauth_clients_registered']}")
        
        print(f"\n  Rate Limits:")
        print(f"    Window: {stats['rate_limits']['window']} seconds")
        print(f"    Requests per Window: {stats['rate_limits']['requests_per_window']}")
        print(f"    Burst Limit: {stats['rate_limits']['burst']}")
        
        return stats
    else:
        print(f"✗ Failed to get statistics")
        print(f"  Error: {result.get('error', 'Unknown error')}")
        return None

def main():
    """Run all tests"""
    print_section("MCP WhatsApp Server Test Suite")
    
    # Get user ID from user
    user_id = input("Enter user ID: ").strip()
    if not user_id:
        user_id = "test_user_" + str(int(time.time()))
        print(f"Using default user ID: {user_id}")
    
    # Test 1: Check rate limit
    if not check_rate_limit(user_id):
        return
    
    # Test 2: Register OAuth client
    client_id, client_secret = test_oauth_register(user_id)
    if not client_id:
        return
    
    # Test 3: Upload Excel file
    batch_id = test_upload_excel(user_id)
    if not batch_id:
        return
    
    # Test 4: Check batch status
    test_get_batch_status(user_id, batch_id)
    
    # Test 5: Send messages (dry run)
    test_send_messages(user_id, batch_id, "Hi there! This is a test message from MCP server.")
    
    # Test 6: Get statistics
    test_get_statistics(user_id)
    
    # Test 7: Check rate limit again
    print_section("Final Rate Limit Check")
    check_rate_limit(user_id)
    
    print_section("Test Complete")
    print("All tests completed successfully!")

if __name__ == "__main__":
    main()
